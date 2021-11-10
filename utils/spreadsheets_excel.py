import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):     # 'transactions.xlsx'
    # Load a workbook and return a workbook object
    wb = xl.load_workbook(filename)
    # Access particular sheet in a workbook
    sheet = wb['Sheet1']

    print(f'\nProcessing {sheet} from \'{wb}\'... \n')

    for row in range(2, sheet.max_row + 1):  # (max_row => max number of rows)
        cell = sheet.cell(row, 3)  # sheet['a1'] is another way to access a cell
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Returns values of column 4 and rows 2 to max_row as an object
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    # Create instance of BarChart and pass 'values' to chart and add chart to sheet at column 'E2'
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save('transactions2.xlsx')  
    print('\nFile Saved.\n')

process_workbook("transactions.xlsx")